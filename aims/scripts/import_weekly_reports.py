#!/usr/bin/env python3
from __future__ import annotations

import argparse
import datetime as dt
import difflib
import json
import math
import re
import sys
import urllib.error
import urllib.parse
import urllib.request
from collections import Counter, defaultdict
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


AIMS_ROOT = Path(__file__).resolve().parents[1]
DEFAULT_DOCS_DIR = AIMS_ROOT / "docs"
DEFAULT_ENV_FILE = AIMS_ROOT / ".env.dev"
DEFAULT_REPORT_FILE = AIMS_ROOT / "docs" / "weekly-report-import-report.json"
DEFAULT_DIRECTORY_URL = "http://localhost:3000/api/v1/directory/users"
DEFAULT_CURRENT_USER = "zhouguangying"

HEADER_ALIASES = {
    "departmentName": ("隶属部门",),
    "projectTypeName": ("项目类型",),
    "projectCode": ("项目编号",),
    "projectName": ("项目名称",),
    "projectManagerName": ("项目经理",),
    "initiationStatus": ("立项情况",),
    "currentStage": ("当前阶段",),
    "progressStatus": ("进度情况",),
    "completionPercent": ("总体完成进度",),
    "contractStatus": ("合同状态",),
    "paymentStatus": ("回款情况",),
    "cumulativeLaborCost": ("累计人力成本",),
    "previousWeekDays": ("上周人力投入",),
    "thisWeekDays": ("本周人力投入",),
    "laborDelta": ("人力较上周",),
    "majorRisks": ("重大问题", "风险"),
    "coordinationNeeds": ("待协调资源",),
    "remarks": ("备注",),
}

USER_NAME_ALIASES = {
    "段幼轩": "段佑轩",
    "韩华斌": "韩化斌",
    "穆德坤": "慕德坤",
    "潘春雷": "潘春蕾",
    "徐建飞": "许建飞",
    "赵文斌": "赵文彬",
}

PROJECT_CODE_ALIASES = {
    "22240822": "22250822",
}

PROJECT_NAME_CODE_ALIASES = {
    "技术支撑": "21250807J",
    "房屋租赁管理汇房云租": "21250925HF",
}

IGNORED_OWNER_NAMES = {"责任人", "项目经理", "合计", "小计", "总计", "无", "暂无"}


@dataclass
class SummaryRecord:
    file: str
    report_year: int
    report_week: int
    project_code: str = ""
    project_name: str = ""
    department_name: str = ""
    project_type_name: str = ""
    project_manager_name: str = ""
    initiation_status: str = ""
    current_stage: str = ""
    progress_status: str = ""
    completion_percent: float | None = None
    contract_status: str = ""
    contract_amount: float | None = None
    payment_status: str = ""
    cumulative_labor_cost: float | None = None
    previous_week_days: float | None = None
    this_week_days: float | None = None
    labor_delta: str = ""
    major_risks: str = ""
    coordination_needs: str = ""
    remarks: str = ""


@dataclass
class DetailGroup:
    file: str
    sheet: str
    report_year: int
    report_week: int
    source_text: str = ""
    project_code: str = ""
    project_name: str = ""
    project_manager_name: str = ""
    current_progress: str = ""
    payment_status: str = ""
    total_days: float | None = None
    member_days: Counter[str] = field(default_factory=Counter)
    work_items: list[dict[str, Any]] = field(default_factory=list)


def read_env(path: Path) -> dict[str, str]:
    values: dict[str, str] = {}
    if not path.exists():
        return values
    for raw_line in path.read_text(encoding="utf-8").splitlines():
        line = raw_line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        values[key.strip()] = value.strip().strip('"').strip("'")
    return values


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    text = str(value).replace("\r\n", "\n").replace("\r", "\n")
    text = re.sub(r"[ \t]+", " ", text)
    return text.strip()


def squash_text(value: Any) -> str:
    return re.sub(r"\s+", " ", clean_text(value)).strip()


def normalize_key(value: Any) -> str:
    text = clean_text(value).lower()
    return re.sub(r"[\s　（）()／/、:：;；,，.。\\-_\[\]【】<>《》]+", "", text)


def normalize_project_name(value: Any) -> str:
    text = normalize_key(value)
    for token in ("有限责任公司", "股份有限公司", "项目", "系统", "平台", "建设", "升级改造"):
        text = text.replace(token, "")
    return text


def normalize_code(value: Any) -> str:
    text = clean_text(value)
    if not text:
        return ""
    text = text.replace(".0", "") if re.fullmatch(r"\d+\.0", text) else text
    text = re.sub(r"\s+", "", text)
    text = text.strip("/／:：-")
    return text.upper()


def parse_number(value: Any) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        if math.isnan(float(value)):
            return None
        return float(value)
    text = clean_text(value)
    if not text or text in {"——", "-", "/", "无", "暂无"}:
        return None
    text = text.replace(",", "")
    match = re.search(r"-?\d+(?:\.\d+)?", text)
    if not match:
        return None
    return float(match.group(0))


def parse_percent(value: Any) -> float | None:
    number = parse_number(value)
    if number is None:
        return None
    text = clean_text(value)
    if isinstance(value, (int, float)) and abs(number) <= 1:
        return round(number * 100, 2)
    if "%" not in text and 0 < number <= 1:
        return round(number * 100, 2)
    return round(number, 2)


def rounded(value: float | None, digits: int = 2) -> float | None:
    if value is None:
        return None
    return round(float(value), digits)


def file_start_date(path: Path) -> dt.date:
    match = re.search(r"(\d{4})\.(\d{1,2})\.(\d{1,2})", path.name)
    if not match:
        raise ValueError(f"Cannot parse start date from {path.name}")
    year, month, day = map(int, match.groups())
    return dt.date(year, month, day)


def week_from_file(path: Path) -> tuple[int, int]:
    start = file_start_date(path)
    if start.weekday() == 6:
        start = start + dt.timedelta(days=1)
    iso = start.isocalendar()
    return iso.year, iso.week


def workbook_files(docs_dir: Path) -> list[Path]:
    files = [p for p in docs_dir.glob("*.xlsx") if not p.name.startswith("~$")]
    return sorted(files, key=file_start_date)


def header_columns(ws) -> tuple[int, dict[int, str]]:
    for row in range(1, min(ws.max_row, 10) + 1):
        values = {col: clean_text(ws.cell(row, col).value) for col in range(1, ws.max_column + 1)}
        normalized = "".join(normalize_key(v) for v in values.values())
        if "项目编号" in normalized and "项目名称" in normalized:
            return row, values
    raise ValueError(f"Cannot find summary header in sheet {ws.title}")


def find_header(headers: dict[int, str], key: str) -> int | None:
    tokens = HEADER_ALIASES[key]
    normalized_tokens = [normalize_key(token) for token in tokens]
    for index, header in headers.items():
        normalized_header = normalize_key(header)
        if all(token in normalized_header for token in normalized_tokens):
            return index
    return None


def cell_by_key(ws, row: int, columns: dict[str, int | None], key: str) -> Any:
    column = columns.get(key)
    if not column:
        return None
    return ws.cell(row, column).value


def parse_contract(value: Any) -> tuple[str, float | None]:
    number = parse_number(value)
    text = clean_text(value)
    if number is not None and text and re.fullmatch(r"[-\d., 万元]+", text):
        return "", number
    if isinstance(value, (int, float)) and number is not None:
        return "", number
    return text if text not in {"——", "-"} else "", None


def parse_summary_sheet(path: Path, wb) -> list[SummaryRecord]:
    if "项目汇总" not in wb.sheetnames:
        return []
    ws = wb["项目汇总"]
    report_year, report_week = week_from_file(path)
    header_row, headers = header_columns(ws)
    columns = {key: find_header(headers, key) for key in HEADER_ALIASES}
    records: list[SummaryRecord] = []
    last_department = ""
    last_type = ""

    for row in range(header_row + 1, ws.max_row + 1):
        project_code = normalize_code(cell_by_key(ws, row, columns, "projectCode"))
        project_name = squash_text(cell_by_key(ws, row, columns, "projectName"))
        if not project_code and not project_name:
            continue

        department_name = clean_text(cell_by_key(ws, row, columns, "departmentName"))
        project_type_name = clean_text(cell_by_key(ws, row, columns, "projectTypeName"))
        if department_name:
            last_department = department_name
        if project_type_name:
            last_type = project_type_name

        contract_status, contract_amount = parse_contract(cell_by_key(ws, row, columns, "contractStatus"))
        records.append(SummaryRecord(
            file=path.name,
            report_year=report_year,
            report_week=report_week,
            project_code=project_code,
            project_name=project_name,
            department_name=last_department,
            project_type_name=last_type,
            project_manager_name=squash_text(cell_by_key(ws, row, columns, "projectManagerName")),
            initiation_status=squash_text(cell_by_key(ws, row, columns, "initiationStatus")),
            current_stage=squash_text(cell_by_key(ws, row, columns, "currentStage")),
            progress_status=squash_text(cell_by_key(ws, row, columns, "progressStatus")),
            completion_percent=parse_percent(cell_by_key(ws, row, columns, "completionPercent")),
            contract_status=contract_status,
            contract_amount=rounded(contract_amount),
            payment_status=squash_text(cell_by_key(ws, row, columns, "paymentStatus")),
            cumulative_labor_cost=rounded(parse_number(cell_by_key(ws, row, columns, "cumulativeLaborCost")), 3),
            previous_week_days=rounded(parse_number(cell_by_key(ws, row, columns, "previousWeekDays")), 3),
            this_week_days=rounded(parse_number(cell_by_key(ws, row, columns, "thisWeekDays")), 3),
            labor_delta=squash_text(cell_by_key(ws, row, columns, "laborDelta")),
            major_risks=clean_text(cell_by_key(ws, row, columns, "majorRisks")),
            coordination_needs=clean_text(cell_by_key(ws, row, columns, "coordinationNeeds")),
            remarks=clean_text(cell_by_key(ws, row, columns, "remarks")),
        ))
    return records


def extract_project_info(value: Any) -> tuple[str, str]:
    text = squash_text(value)
    if not text or normalize_key(text) == "项目信息":
        return "", ""
    code = ""
    code_match = re.search(r"项目编号\s*[:：]?\s*[\/／]?\s*([A-Za-z0-9_.-]+)", text, re.I)
    if code_match:
        code = normalize_code(code_match.group(1))
    name = re.split(r"项目编号", text, maxsplit=1)[0]
    name = re.sub(r"^项目名称\s*[:：]?", "", name).strip()
    name = name.strip(" /／:：-")
    if not name and text and not code_match:
        name = text.strip(" /／:：-")
    return code, name


def is_headerish_row(*values: Any) -> bool:
    text = normalize_key("".join(clean_text(v) for v in values))
    return any(marker in text for marker in ("工作计划", "任务简述", "责任人", "工作量人日"))


def append_work_item(group: DetailGroup, row_number: int, plan_type: str, module_name: str, task_summary: str,
                     owner_name: str, completion: float | None, incomplete_reason: str, workload_days: float | None) -> None:
    if not task_summary:
        return
    group.work_items.append({
        "planType": plan_type,
        "sourceType": "manual",
        "moduleName": module_name,
        "sortOrder": len(group.work_items) + 1,
        "taskSummary": task_summary,
        "ownerName": owner_name,
        "completionPercent": completion,
        "incompleteReason": incomplete_reason,
        "workloadDays": rounded(workload_days),
        "_sourceRow": row_number,
        "_sourceSheet": group.sheet,
    })


def parse_detail_sheet(path: Path, wb) -> list[DetailGroup]:
    report_year, report_week = week_from_file(path)
    groups: list[DetailGroup] = []

    for sheet_name in wb.sheetnames:
        if sheet_name == "项目汇总":
            continue
        ws = wb[sheet_name]
        current: DetailGroup | None = None
        mode = "this_week"
        last_module = ""
        last_task = ""

        def finalize() -> None:
            nonlocal current
            if current and (current.project_code or current.project_name or current.member_days or current.work_items):
                groups.append(current)
            current = None

        for row in range(1, ws.max_row + 1):
            project_cell = ws.cell(row, 1).value
            project_code, project_name = extract_project_info(project_cell)
            if project_code or project_name:
                finalize()
                current = DetailGroup(
                    file=path.name,
                    sheet=sheet_name,
                    report_year=report_year,
                    report_week=report_week,
                    source_text=squash_text(project_cell),
                    project_code=project_code,
                    project_name=project_name,
                    project_manager_name=squash_text(ws.cell(row, 2).value),
                    current_progress=clean_text(ws.cell(row, 3).value),
                    payment_status=clean_text(ws.cell(row, 4).value),
                )
                mode = "this_week"
                last_module = ""
                last_task = ""

            if not current:
                continue

            marker = normalize_key(ws.cell(row, 5).value)
            if "下周计划" in marker:
                mode = "next_week"
                continue
            if "本周工作" in marker:
                mode = "this_week"
            if marker in {"合计", "小计", "总计"}:
                total_days = parse_number(ws.cell(row, 12).value)
                if total_days is not None:
                    current.total_days = rounded(total_days, 3)
                continue

            module_name = squash_text(ws.cell(row, 6).value) or last_module
            task_summary = clean_text(ws.cell(row, 8).value)
            owner_name = squash_text(ws.cell(row, 9).value)
            completion = parse_percent(ws.cell(row, 10).value)
            incomplete_reason = clean_text(ws.cell(row, 11).value)
            workload_days = parse_number(ws.cell(row, 12).value)

            if is_headerish_row(module_name, task_summary, owner_name):
                continue
            if module_name:
                last_module = module_name
            if task_summary:
                last_task = task_summary
            elif owner_name or workload_days is not None:
                task_summary = last_task

            if owner_name and owner_name not in IGNORED_OWNER_NAMES and workload_days is not None and mode == "this_week":
                current.member_days[owner_name] += float(workload_days)

            has_detail = bool(task_summary or owner_name or incomplete_reason or workload_days is not None)
            if has_detail:
                append_work_item(
                    current,
                    row,
                    mode,
                    module_name,
                    task_summary,
                    owner_name if owner_name not in IGNORED_OWNER_NAMES else "",
                    completion,
                    incomplete_reason,
                    workload_days,
                )

        finalize()
    return groups


def request_json(method: str, url: str, token: str | None = None, body: Any | None = None, timeout: int = 30) -> Any:
    headers: dict[str, str] = {"Accept": "application/json"}
    payload = None
    if token:
        headers["Authorization"] = f"Bearer {token}"
    if body is not None:
        payload = json.dumps(body, ensure_ascii=False).encode("utf-8")
        headers["Content-Type"] = "application/json"
    req = urllib.request.Request(url, data=payload, headers=headers, method=method)
    try:
        with urllib.request.urlopen(req, timeout=timeout) as response:
            text = response.read().decode("utf-8")
            return json.loads(text) if text else None
    except urllib.error.HTTPError as error:
        detail = error.read().decode("utf-8", "replace")
        raise RuntimeError(f"{method} {url} failed: HTTP {error.code} {detail}") from error


def data_items(response: Any) -> list[dict[str, Any]]:
    if isinstance(response, dict):
        data = response.get("data", response)
        if isinstance(data, dict) and isinstance(data.get("items"), list):
            return data["items"]
        if isinstance(data, list):
            return data
    return []


def fetch_runtime_projects(runtime_url: str, token: str, current_user: str) -> list[dict[str, Any]]:
    params = {
        "current_user": current_user,
        "current_user_is_project_admin": "1",
        "pageSize": "1000",
    }
    url = f"{runtime_url.rstrip('/')}/v1/aims/admin/projects?{urllib.parse.urlencode(params)}"
    return data_items(request_json("GET", url, token=token))


def fetch_project_members(runtime_url: str, token: str, current_user: str, project_id: int) -> list[dict[str, Any]]:
    params = {
        "current_user": current_user,
        "current_user_is_project_admin": "1",
        "pageSize": "1000",
    }
    url = f"{runtime_url.rstrip('/')}/v1/aims/projects/{project_id}/members?{urllib.parse.urlencode(params)}"
    return data_items(request_json("GET", url, token=token))


def fetch_directory_users(directory_url: str) -> list[dict[str, Any]]:
    users: list[dict[str, Any]] = []
    page = 1
    page_size = 1000
    while True:
        separator = "&" if "?" in directory_url else "?"
        url = f"{directory_url}{separator}{urllib.parse.urlencode({'page': page, 'pageSize': page_size})}"
        response = request_json("GET", url, token=None, timeout=20)
        items = data_items(response)
        users.extend(items)
        data = response.get("data", {}) if isinstance(response, dict) else {}
        total = int(data.get("total") or len(users))
        if not items or len(users) >= total:
            break
        page += 1
    dedup: dict[str, dict[str, Any]] = {}
    for user in users:
        uid = clean_text(user.get("uid"))
        if uid:
            dedup[uid] = user
    return list(dedup.values())


def user_display_name(user: dict[str, Any]) -> str:
    return clean_text(user.get("realName")) or clean_text(user.get("displayName")) or clean_text(user.get("name")) or clean_text(user.get("uid"))


def build_user_index(users: list[dict[str, Any]]) -> tuple[dict[str, list[dict[str, Any]]], dict[str, dict[str, Any]]]:
    by_name: dict[str, list[dict[str, Any]]] = defaultdict(list)
    by_uid: dict[str, dict[str, Any]] = {}
    for user in users:
        uid = clean_text(user.get("uid"))
        if not uid:
            continue
        by_uid[uid] = user
        names = {
            clean_text(user.get("realName")),
            clean_text(user.get("displayName")),
            clean_text(user.get("name")),
            clean_text(user.get("username")),
        }
        for name in names:
            key = normalize_key(name)
            if key:
                by_name[key].append(user)
    return by_name, by_uid


def resolve_user(name: str, by_name: dict[str, list[dict[str, Any]]]) -> tuple[str, str, str]:
    normalized = normalize_key(USER_NAME_ALIASES.get(name, name))
    if not normalized:
        return "", "", "empty"
    candidates = by_name.get(normalized, [])
    if len(candidates) == 1:
        user = candidates[0]
        return clean_text(user.get("uid")), user_display_name(user), "exact"
    if len(candidates) > 1:
        return "", "", "ambiguous_user"

    best_key = ""
    best_score = 0.0
    for key in by_name:
        score = difflib.SequenceMatcher(None, normalized, key).ratio()
        if score > best_score:
            best_key = key
            best_score = score
    if best_score >= 0.86 and len(by_name[best_key]) == 1:
        user = by_name[best_key][0]
        return clean_text(user.get("uid")), user_display_name(user), f"fuzzy:{best_score:.2f}"
    return "", "", "user_not_found"


def build_project_index(projects: list[dict[str, Any]]) -> tuple[dict[str, list[dict[str, Any]]], list[dict[str, Any]]]:
    by_code: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for project in projects:
        for key in ("internal_code", "project_code", "short_name"):
            code = normalize_code(project.get(key))
            if code:
                by_code[code].append(project)
    return by_code, projects


def project_identity(project: dict[str, Any]) -> str:
    return f"{project.get('id')} {project.get('internal_code') or project.get('project_code') or ''} {project.get('name')}"


def resolve_project(project_code: str, project_name: str, by_code: dict[str, list[dict[str, Any]]],
                    projects: list[dict[str, Any]]) -> tuple[dict[str, Any] | None, str, float, str]:
    code = PROJECT_CODE_ALIASES.get(normalize_code(project_code), normalize_code(project_code))
    normalized_project_name = normalize_project_name(project_name)
    alias_code = PROJECT_NAME_CODE_ALIASES.get(normalized_project_name)
    if not alias_code:
        for alias_name, target_code in PROJECT_NAME_CODE_ALIASES.items():
            if normalized_project_name.startswith(alias_name):
                alias_code = target_code
                break
    if alias_code:
        code = normalize_code(alias_code)
    if code and by_code.get(code):
        candidates = by_code[code]
        if len(candidates) == 1:
            return candidates[0], "name_alias" if alias_code else "code", 1.0, ""
        name_norm = normalize_project_name(project_name)
        best = max(candidates, key=lambda item: difflib.SequenceMatcher(None, name_norm, normalize_project_name(item.get("name"))).ratio())
        score = difflib.SequenceMatcher(None, name_norm, normalize_project_name(best.get("name"))).ratio()
        return best, "code_ambiguous_name_best", score, ""

    source = normalize_project_name(project_name)
    if not source:
        return None, "missing_project_name", 0.0, ""
    best_project: dict[str, Any] | None = None
    best_score = 0.0
    for project in projects:
        names = [project.get("name"), project.get("short_name"), project.get("project_code"), project.get("internal_code")]
        score = max((project_name_score(source, normalize_project_name(name)) for name in names if clean_text(name)), default=0.0)
        if score > best_score:
            best_project = project
            best_score = score
    if best_project and best_score >= 0.62:
        return best_project, "fuzzy_name", best_score, ""
    return None, "project_not_found", best_score, project_identity(best_project) if best_project else ""


def project_name_score(source: str, target: str) -> float:
    if not source or not target:
        return 0.0
    if source == target:
        return 1.0
    if source in target or target in source:
        shorter = min(len(source), len(target))
        longer = max(len(source), len(target))
        return 0.72 + 0.25 * (shorter / longer)
    return difflib.SequenceMatcher(None, source, target).ratio()


def summary_to_payload(record: SummaryRecord) -> dict[str, Any]:
    payload = {
        "mainWork": "",
        "overallProgress": record.remarks or record.progress_status,
        "departmentName": record.department_name,
        "projectTypeName": record.project_type_name,
        "projectManagerName": record.project_manager_name,
        "initiationStatus": record.initiation_status,
        "currentStage": record.current_stage,
        "progressStatus": record.progress_status,
        "completionPercent": record.completion_percent,
        "contractStatus": record.contract_status,
        "contractAmount": record.contract_amount,
        "paymentStatus": record.payment_status,
        "cumulativeLaborCost": record.cumulative_labor_cost,
        "majorRisks": record.major_risks,
        "coordinationNeeds": record.coordination_needs,
        "remarks": record.remarks,
    }
    return {key: value for key, value in payload.items() if value not in ("", None)}


def build_main_work(work_items: list[dict[str, Any]]) -> str:
    lines: list[str] = []
    seen: set[str] = set()
    for item in work_items:
        if item.get("planType") != "this_week":
            continue
        summary = clean_text(item.get("taskSummary"))
        if not summary:
            continue
        owner = clean_text(item.get("ownerName"))
        days = item.get("workloadDays")
        suffix_parts = []
        if owner:
            suffix_parts.append(owner)
        if days is not None:
            suffix_parts.append(f"{days:g}人日")
        line = f"- {summary}"
        if suffix_parts:
            line += f"（{'，'.join(suffix_parts)}）"
        key = normalize_key(line)
        if key in seen:
            continue
        seen.add(key)
        lines.append(line)
        if len(lines) >= 30:
            break
    return "\n".join(lines)


def prepare_import(docs_dir: Path, runtime_url: str, runtime_token: str, directory_url: str,
                   current_user: str, auto_add_project_members: bool = False) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    projects = fetch_runtime_projects(runtime_url, runtime_token, current_user)
    project_by_code, project_list = build_project_index(projects)

    users = fetch_directory_users(directory_url)
    users_by_name, users_by_uid = build_user_index(users)

    all_member_sets: dict[int, set[str]] = {}
    all_project_members: dict[int, list[dict[str, Any]]] = {}
    for project in projects:
        project_id = int(project["id"])
        try:
            members = fetch_project_members(runtime_url, runtime_token, current_user, project_id)
        except Exception as error:
            members = []
            print(f"WARN: failed to fetch members for project {project_id}: {error}", file=sys.stderr)
        all_project_members[project_id] = members
        member_uids = {clean_text(member.get("uid")) for member in members if clean_text(member.get("uid")) and clean_text(member.get("status")) != "suspended"}
        leader_uid = clean_text(project.get("leader_uid"))
        if leader_uid:
            member_uids.add(leader_uid)
        all_member_sets[project_id] = member_uids

    report: dict[str, Any] = {
        "mode": "preview",
        "generatedAt": dt.datetime.now().isoformat(timespec="seconds"),
        "files": [],
        "stats": {
            "files": 0,
            "summaryRows": 0,
            "detailGroups": 0,
            "matchedReports": 0,
            "readyReports": 0,
            "readyEntries": 0,
            "readyWorkItems": 0,
        },
        "unmatchedProjects": [],
        "unmatchedMembers": [],
        "membersToAdd": [],
        "skippedReports": [],
        "emptyReportsIgnored": [],
        "workloadMismatches": [],
        "preparedReports": [],
    }
    prepared: list[dict[str, Any]] = []
    members_to_add: dict[tuple[int, str], dict[str, Any]] = {}

    for path in workbook_files(docs_dir):
        report_year, report_week = week_from_file(path)
        wb = load_workbook(path, data_only=True, read_only=True)
        summaries = parse_summary_sheet(path, wb)
        details = parse_detail_sheet(path, wb)
        report["files"].append({
            "file": path.name,
            "reportYear": report_year,
            "reportWeek": report_week,
            "summaryRows": len(summaries),
            "detailGroups": len(details),
        })
        report["stats"]["files"] += 1
        report["stats"]["summaryRows"] += len(summaries)
        report["stats"]["detailGroups"] += len(details)

        records_by_project: dict[int, dict[str, Any]] = {}

        def ensure_project_record(project: dict[str, Any], source: dict[str, Any]) -> dict[str, Any]:
            project_id = int(project["id"])
            if project_id not in records_by_project:
                records_by_project[project_id] = {
                    "file": path.name,
                    "reportYear": report_year,
                    "reportWeek": report_week,
                    "projectId": project_id,
                    "projectName": project.get("name"),
                    "projectCode": project.get("project_code"),
                    "internalCode": project.get("internal_code"),
                    "summary": None,
                    "summaryPayload": {},
                    "summaryThisWeekDays": None,
                    "memberDays": Counter(),
                    "workItems": [],
                    "sources": [],
                    "matchSources": [],
                    "project": project,
                }
            records_by_project[project_id]["sources"].append(source)
            return records_by_project[project_id]

        for summary in summaries:
            project, method, score, best = resolve_project(summary.project_code, summary.project_name, project_by_code, project_list)
            if not project:
                report["unmatchedProjects"].append({
                    "file": path.name,
                    "reportYear": report_year,
                    "reportWeek": report_week,
                    "source": "summary",
                    "projectCode": summary.project_code,
                    "projectName": summary.project_name,
                    "reason": method,
                    "bestScore": round(score, 3),
                    "bestCandidate": best,
                })
                continue
            record = ensure_project_record(project, {
                "type": "summary",
                "projectCode": summary.project_code,
                "projectName": summary.project_name,
                "matchMethod": method,
                "matchScore": round(score, 3),
            })
            record["summary"] = summary
            record["summaryPayload"] = summary_to_payload(summary)
            record["summaryThisWeekDays"] = summary.this_week_days
            record["matchSources"].append(method)

        for group in details:
            project, method, score, best = resolve_project(group.project_code, group.project_name, project_by_code, project_list)
            if not project:
                report["unmatchedProjects"].append({
                    "file": path.name,
                    "reportYear": report_year,
                    "reportWeek": report_week,
                    "source": "detail",
                    "sheet": group.sheet,
                    "projectCode": group.project_code,
                    "projectName": group.project_name,
                    "reason": method,
                    "bestScore": round(score, 3),
                    "bestCandidate": best,
                })
                continue
            record = ensure_project_record(project, {
                "type": "detail",
                "sheet": group.sheet,
                "projectCode": group.project_code,
                "projectName": group.project_name,
                "matchMethod": method,
                "matchScore": round(score, 3),
            })
            record["memberDays"].update(group.member_days)
            record["workItems"].extend(group.work_items)
            detail_sum = round(sum(group.member_days.values()), 3)
            if group.total_days is not None and abs(detail_sum - group.total_days) > 0.11:
                report["workloadMismatches"].append({
                    "file": path.name,
                    "reportYear": report_year,
                    "reportWeek": report_week,
                    "projectId": int(project["id"]),
                    "projectName": project.get("name"),
                    "source": "detail_total",
                    "sheet": group.sheet,
                    "detailMemberDays": detail_sum,
                    "detailTotalDays": group.total_days,
                })

        for project_id, record in sorted(records_by_project.items(), key=lambda item: item[0]):
            report["stats"]["matchedReports"] += 1
            member_set = all_member_sets.get(project_id, set())
            entries: list[dict[str, Any]] = []
            resolved_member_days = 0.0
            for owner_name, days in sorted(record["memberDays"].items(), key=lambda item: normalize_key(item[0])):
                if days <= 0:
                    continue
                uid, display_name, user_method = resolve_user(owner_name, users_by_name)
                if not uid:
                    report["unmatchedMembers"].append({
                        "file": path.name,
                        "reportYear": report_year,
                        "reportWeek": report_week,
                        "projectId": project_id,
                        "projectName": record["projectName"],
                        "memberName": owner_name,
                        "days": round(days, 3),
                        "reason": user_method,
                    })
                    continue
                if uid not in member_set:
                    if auto_add_project_members:
                        members_to_add.setdefault((project_id, uid), {
                            "projectId": project_id,
                            "projectName": record["projectName"],
                            "uid": uid,
                            "realName": display_name or owner_name,
                            "role": "member",
                            "status": "active",
                            "source": "weekly_report_import",
                        })
                        member_set.add(uid)
                    else:
                        report["unmatchedMembers"].append({
                            "file": path.name,
                            "reportYear": report_year,
                            "reportWeek": report_week,
                            "projectId": project_id,
                            "projectName": record["projectName"],
                            "memberName": owner_name,
                            "uid": uid,
                            "days": round(days, 3),
                            "reason": "not_active_project_member",
                        })
                        continue
                hours = round(days * 8, 2)
                if hours > 168:
                    report["unmatchedMembers"].append({
                        "file": path.name,
                        "reportYear": report_year,
                        "reportWeek": report_week,
                        "projectId": project_id,
                        "projectName": record["projectName"],
                        "memberName": owner_name,
                        "uid": uid,
                        "days": round(days, 3),
                        "reason": "hours_exceed_api_limit",
                    })
                    continue
                entries.append({
                    "uid": uid,
                    "hours": hours,
                    "allocationPercent": round(hours / 40 * 100, 2),
                    "_memberName": display_name or owner_name,
                    "_days": round(days, 3),
                    "_matchMethod": user_method,
                })
                resolved_member_days += days

            for item in record["workItems"]:
                owner_name = clean_text(item.get("ownerName"))
                if owner_name:
                    uid, display_name, _ = resolve_user(owner_name, users_by_name)
                    if uid and uid in member_set:
                        item["ownerUid"] = uid
                        item["ownerName"] = display_name or owner_name
                for private_key in [key for key in item if key.startswith("_")]:
                    item.pop(private_key, None)

            detail_days = round(sum(record["memberDays"].values()), 3)
            summary_days = record.get("summaryThisWeekDays")
            if summary_days is not None and detail_days and abs(detail_days - float(summary_days)) > 0.26:
                report["workloadMismatches"].append({
                    "file": path.name,
                    "reportYear": report_year,
                    "reportWeek": report_week,
                    "projectId": project_id,
                    "projectName": record["projectName"],
                    "source": "summary_vs_detail",
                    "summaryThisWeekDays": summary_days,
                    "detailMemberDays": detail_days,
                })
            if summary_days is not None and abs(resolved_member_days - float(summary_days)) > 0.26:
                report["workloadMismatches"].append({
                    "file": path.name,
                    "reportYear": report_year,
                    "reportWeek": report_week,
                    "projectId": project_id,
                    "projectName": record["projectName"],
                    "source": "summary_vs_resolved_entries",
                    "summaryThisWeekDays": summary_days,
                    "resolvedEntryDays": round(resolved_member_days, 3),
                })

            if not entries:
                skipped_item = {
                    "file": path.name,
                    "reportYear": report_year,
                    "reportWeek": report_week,
                    "projectId": project_id,
                    "projectName": record["projectName"],
                    "reason": "no_resolved_project_member_entries",
                    "summaryThisWeekDays": summary_days,
                    "detailMemberDays": detail_days,
                }
                if (summary_days is None or abs(float(summary_days)) <= 0.001) and abs(detail_days) <= 0.001:
                    report["emptyReportsIgnored"].append(skipped_item)
                else:
                    report["skippedReports"].append(skipped_item)
                continue

            work_items = record["workItems"]
            payload = {
                "reportYear": report_year,
                "reportWeek": report_week,
                "status": "submitted",
                **record.get("summaryPayload", {}),
                "entries": [{key: value for key, value in entry.items() if not key.startswith("_")} for entry in entries],
                "workItems": work_items,
            }
            if not payload.get("mainWork"):
                main_work = build_main_work(work_items)
                if main_work:
                    payload["mainWork"] = main_work

            prepared_record = {
                "file": path.name,
                "reportYear": report_year,
                "reportWeek": report_week,
                "projectId": project_id,
                "projectName": record["projectName"],
                "projectCode": record["projectCode"],
                "internalCode": record["internalCode"],
                "summaryThisWeekDays": summary_days,
                "detailMemberDays": detail_days,
                "entryDays": round(sum(entry["_days"] for entry in entries), 3),
                "entryCount": len(entries),
                "workItemCount": len(work_items),
                "matchSources": sorted(set(record["matchSources"])),
                "payload": payload,
            }
            prepared.append(prepared_record)
            report["preparedReports"].append({key: value for key, value in prepared_record.items() if key != "payload"})
            report["stats"]["readyReports"] += 1
            report["stats"]["readyEntries"] += len(entries)
            report["stats"]["readyWorkItems"] += len(work_items)

    report["membersToAdd"] = sorted(members_to_add.values(), key=lambda item: (item["projectId"], item["uid"]))
    report["stats"]["membersToAdd"] = len(report["membersToAdd"])
    return prepared, report


def apply_import(prepared: list[dict[str, Any]], report: dict[str, Any], runtime_url: str,
                 runtime_token: str, current_user: str) -> None:
    report["mode"] = "apply"
    report["appliedReports"] = []
    report["failedReports"] = []
    for item in prepared:
        params = {
            "current_user": current_user,
            "current_user_is_project_admin": "1",
        }
        url = f"{runtime_url.rstrip('/')}/v1/aims/projects/{item['projectId']}/weekly-reports?{urllib.parse.urlencode(params)}"
        try:
            response = request_json("POST", url, token=runtime_token, body=item["payload"], timeout=40)
            data = response.get("data", {}) if isinstance(response, dict) else {}
            report["appliedReports"].append({
                "file": item["file"],
                "reportYear": item["reportYear"],
                "reportWeek": item["reportWeek"],
                "projectId": item["projectId"],
                "projectName": item["projectName"],
                "reportId": data.get("id"),
                "totalHours": data.get("totalHours"),
                "memberCount": data.get("memberCount"),
                "workItemCount": item["workItemCount"],
            })
        except Exception as error:
            report["failedReports"].append({
                "file": item["file"],
                "reportYear": item["reportYear"],
                "reportWeek": item["reportWeek"],
                "projectId": item["projectId"],
                "projectName": item["projectName"],
                "error": str(error),
            })


def write_report(path: Path, report: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")


def main() -> int:
    parser = argparse.ArgumentParser(description="Import AIMS weekly report Excel files.")
    parser.add_argument("--docs-dir", type=Path, default=DEFAULT_DOCS_DIR)
    parser.add_argument("--env-file", type=Path, default=DEFAULT_ENV_FILE)
    parser.add_argument("--directory-url", default=DEFAULT_DIRECTORY_URL)
    parser.add_argument("--runtime-url", default="")
    parser.add_argument("--current-user", default=DEFAULT_CURRENT_USER)
    parser.add_argument("--report", type=Path, default=DEFAULT_REPORT_FILE)
    parser.add_argument("--apply", action="store_true", help="Write prepared records to data-runtime.")
    parser.add_argument("--include-payloads", action="store_true", help="Include prepared write payloads in the JSON report.")
    parser.add_argument("--auto-add-project-members", action="store_true", help="Treat existing directory users as project members to add before import.")
    args = parser.parse_args()

    env = read_env(args.env_file)
    runtime_url = args.runtime_url.strip() or env.get("HZY_TENANT_RUNTIME_URL", "").strip()
    runtime_token = (env.get("HZY_TENANT_RUNTIME_TOKEN") or env.get("HZY_DATA_RUNTIME_STATIC_TOKEN") or "").strip()
    if not runtime_url or not runtime_token:
        print("HZY_TENANT_RUNTIME_URL and HZY_TENANT_RUNTIME_TOKEN are required in env file.", file=sys.stderr)
        return 2

    prepared, report = prepare_import(
        args.docs_dir,
        runtime_url,
        runtime_token,
        args.directory_url,
        args.current_user,
        auto_add_project_members=args.auto_add_project_members,
    )
    if args.include_payloads:
        report["preparedPayloads"] = prepared
    if args.apply:
        apply_import(prepared, report, runtime_url, runtime_token, args.current_user)

    write_report(args.report, report)
    stats = report["stats"]
    print(json.dumps({
        "mode": report["mode"],
        "files": stats["files"],
        "summaryRows": stats["summaryRows"],
        "detailGroups": stats["detailGroups"],
        "readyReports": stats["readyReports"],
        "readyEntries": stats["readyEntries"],
        "readyWorkItems": stats["readyWorkItems"],
        "membersToAdd": len(report.get("membersToAdd", [])),
        "unmatchedProjects": len(report["unmatchedProjects"]),
        "unmatchedMembers": len(report["unmatchedMembers"]),
        "skippedReports": len(report["skippedReports"]),
        "emptyReportsIgnored": len(report["emptyReportsIgnored"]),
        "workloadMismatches": len(report["workloadMismatches"]),
        "appliedReports": len(report.get("appliedReports", [])),
        "failedReports": len(report.get("failedReports", [])),
        "report": str(args.report),
    }, ensure_ascii=False, indent=2))
    return 1 if report.get("failedReports") else 0


if __name__ == "__main__":
    raise SystemExit(main())
